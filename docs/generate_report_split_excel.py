# -*- coding: utf-8 -*-
"""StoryBricks HCI/VR 实验报告分工 Excel — 按策划→设计→功能→测试→总结，尽量不重复。"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = r"d:\Game\StoryBricks\docs\StoryBricks_HCI_VR_ReportSplit.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
PHASE_FILLS = {
    "策划": PatternFill("solid", fgColor="E7E6E6"),
    "设计": PatternFill("solid", fgColor="FCE4D6"),
    "功能": PatternFill("solid", fgColor="E2EFDA"),
    "测试": PatternFill("solid", fgColor="DDEBF7"),
    "总结": PatternFill("solid", fgColor="FFF2CC"),
}
HCI_FILL = PatternFill("solid", fgColor="C6E0B4")
VR_FILL = PatternFill("solid", fgColor="BDD7EE")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="B4B4B4")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_phase_sheet(ws, title, headers, rows):
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    style_header(ws, row=3)
    for i, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = WRAP
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if c == 2:
                if val == "HCI":
                    cell.fill = HCI_FILL
                elif val == "VR":
                    cell.fill = VR_FILL
            if c == 1 and val in PHASE_FILLS:
                cell.fill = PHASE_FILLS[val]


def main():
    wb = Workbook()

    # ═══════════════════════════════════════
    # Sheet 1：总览 + 防重复原则
    # ═══════════════════════════════════════
    ws = wb.active
    ws.title = "0-总览与原则"
    ws.append(["StoryBricks 双课程实验报告分工"])
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for row in [
        ["", ""],
        ["划分维度", "按项目阶段：策划 → 设计 → 功能 → 测试 → 总结"],
        ["HCI 课定位", "研究问题：儿童能不能顺畅地完成故事创作？（体验与交互）"],
        ["VR 课定位", "研究问题：物理积木如何变成虚拟视觉内容并 360° 呈现？（技术与空间）"],
        ["", ""],
        ["防重复原则", ""],
        ["1", "同一功能只在一门课的「功能」表里出现一次"],
        ["2", "HCI 写「用户做什么、看到什么、感受如何」；VR 写「系统怎么检测/生成/渲染」"],
        ["3", "创作页：HCI 写流程与语音；VR 写 ArUco/AR/生图/全景（不写按钮布局）"],
        ["4", "阅读页：HCI 写翻页/录音/360 按钮；VR 写全景文件管线与播放器"],
        ["5", "引言各写 1 段项目背景即可，不复制大段文字；用互引代替重复描述"],
        ["6", "测试完全分开：HCI 测创作可用性；VR 测识别率/生成率/360 沉浸感"],
        ["", ""],
        ["不做内容", "移动 VR 剧场、立体分屏、头显 OpenXR（本期搁置）"],
        ["沉浸阅读", "仅 360° 全景环视（GyroPanorama360Player）"],
    ]:
        ws.append(row)
    set_col_widths(ws, [14, 70])

    # ═══════════════════════════════════════
    # Sheet 2：策划
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("1-策划")
    plan_rows = [
        ["策划", "HCI", "项目立项与用户需求", "明确目标用户（6~10 岁儿童）、使用场景（家庭/课堂）、核心任务（搭积木创作绘本）", "—", "用户画像、场景描述、任务清单", "待写"],
        ["策划", "HCI", "竞品与参考分析", "绘本 App、积木玩具、儿童语音助手、故事创作类产品的交互模式", "—", "竞品对比表（交互维度）", "待写"],
        ["策划", "HCI", "创作流程需求", "完整链路：选故事→读前情→搭积木→摆场景→说话→生成→保存→回看", "—", "用户旅程图（初版）", "待写"],
        ["策划", "HCI", "交互需求规格", "语音优先、步骤清晰、即时反馈、可重来、儿童可理解的语言", "—", "需求列表", "待写"],
        ["策划", "VR", "技术可行性分析", "WebCam+ArUco 能否稳定识别积木、云端 AI 生图延迟、360 全景生成是否可达", "—", "可行性结论", "待写"],
        ["策划", "VR", "虚实融合方案选型", "ArUco 标记体系（ID 1~20）、摄像头俯拍、2D 贴纸叠加（非头显 AR）", "—", "标记分配表（角色/道具）", "待写"],
        ["策划", "VR", "虚拟内容规格", "绘本页 16:9、角色 img2img 参考图、360 全景 2:1 equirectangular 1536×768", "—", "内容规格表", "待写"],
        ["策划", "VR", "技术栈确定", "Unity 2022.3、OpenCV for Unity、Node 生图网关 wan2.6、陀螺仪 360 播放器", "—", "技术栈清单", "待写"],
    ]
    write_phase_sheet(ws2, "阶段一：策划", ["阶段", "课程", "工作包", "具体内容（只写本课）", "关键产出物", "报告素材", "状态"], plan_rows)
    set_col_widths(ws2, [8, 8, 22, 42, 22, 20, 8])

    # ═══════════════════════════════════════
    # Sheet 3：设计
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("2-设计")
    design_rows = [
        ["设计", "HCI", "信息架构", "三层导航：故事库 / 积木作品集 / 我的故事；场景跳转关系", "StoryFlowScenes", "IA 图、场景流程图", "待写"],
        ["设计", "HCI", "创作页交互状态机", "Guide→Building→Capturing→Voice→Confirm→Generating→Done 各阶段用户操作与反馈", "StoryCreationPageBootstrap", "状态机图、线框图", "待写"],
        ["设计", "HCI", "语音助手乐乐", "唤醒词「你好乐乐」、自由叙述、AI 追问脚本结构、TTS 播报时机", "StoryCreationVoiceGateway", "对话样例、话术表", "待写"],
        ["设计", "HCI", "教程交互", "分步滑动、乐乐面板、语音导师问答、3D 预览入口（只设计入口与操作）", "TutorialStepsPageBootstrap", "教程线框图", "待写"],
        ["设计", "HCI", "界面视觉", "StoryCard 卡片、绘本水彩风、儿童配色、TMP 字体、简化 UI 模式", "StoryCard.prefab 等", "UI 截图+标注", "待写"],
        ["设计", "HCI", "阅读页交互", "翻页、故事文本面板、录音/播放/重录、「360° 全景」按钮位置与无资源提示", "CompletedStoryViewerRoot", "阅读页线框图", "已更新"],
        ["设计", "VR", "ArUco 标记设计", "角色 marker ID 分配、检测距离/角度约束、与 StoryDefinition 绑定", "StoryMarkerTaxonomy", "标记对照表", "待写"],
        ["设计", "VR", "AR 叠加方案", "摄像头画面→检测→2D 贴纸坐标映射→角色到场/移动/到齐事件", "StoryCreationArDirector", "AR 数据流图", "待写"],
        ["设计", "VR", "空间语义设计", "marker 像素位置→前后左右关系→驱动 AI 追问内容", "StoryCreationGapAnalyzer", "布局推断规则说明", "待写"],
        ["设计", "VR", "3D 预览方案", "RenderTexture 1024、独立相机、Pointer 拖拽旋转", "TutorialPreview3DOverlay", "渲染管线简图", "待写"],
        ["设计", "VR", "AI 生图管线", "img2img prompt 结构、角色参考图、P2+ 追加锚图、网关 API", "LocalImageGenClient", "API 时序图", "待写"],
        ["设计", "VR", "360° 全景管线", "平面页→/generate-panorama→page_XX_panorama 保存→阅读加载", "CompletedStoryStore", "全景管线流程图", "待完善"],
    ]
    write_phase_sheet(ws3, "阶段二：设计", ["阶段", "课程", "工作包", "具体内容（只写本课）", "关键脚本/资产", "报告素材", "状态"], design_rows)
    set_col_widths(ws3, [8, 8, 22, 42, 24, 20, 8])

    # ═══════════════════════════════════════
    # Sheet 4：功能（核心，每项只出现一次）
    # ═══════════════════════════════════════
    ws4 = wb.create_sheet("3-功能")
    func_rows = [
        # ── HCI 功能（用户可见交互）──
        ["功能", "HCI", "F-H01 故事库", "多故事卡片列表、选故事进入绘本", "StorySummary / StoryCardView", "已完成"],
        ["功能", "HCI", "F-H02 绘本前情", "滑动翻页阅读故事背景、进入搭建", "StoryProloguePictureBook", "已完成"],
        ["功能", "HCI", "F-H03 积木作品集", "本故事关联作品入口、进入教程", "StoryWorks / BrickPortfolioRoot", "已完成"],
        ["功能", "HCI", "F-H04 搭建教程", "分步图文、滑动切换、乐乐 Lottie 吉祥物", "TutorialStepsPageBootstrap", "已完成"],
        ["功能", "HCI", "F-H05 教程语音导师", "搭建过程中向乐乐提问、AI 回答", "TutorialVoiceTutorController", "已完成"],
        ["功能", "HCI", "F-H06 创作页流程", "阶段引导、确认摆好、重新摆、重新生成", "StoryCreationPageBootstrap", "已完成"],
        ["功能", "HCI", "F-H07 创作页语音", "唤醒乐乐、自由叙述、AI 追问、TTS 播放", "StoryCreationVoiceGateway", "已完成"],
        ["功能", "HCI", "F-H08 创作页 UI", "背景、引导文案、操作按钮、摄像头小窗展开", "StoryCreationPageUiBuilder", "已完成"],
        ["功能", "HCI", "F-H09 我的故事库", "已保存绘本列表、选择进入阅读", "CompletedStoryLibrary", "已完成"],
        ["功能", "HCI", "F-H10 平面阅读", "翻页、页码指示、故事文本面板、收起/展开", "CompletedStoryViewerRoot", "已完成"],
        ["功能", "HCI", "F-H11 页内录音", "每页录音、播放、重录", "CompletedStoryPageVoiceRecorder", "已完成"],
        ["功能", "HCI", "F-H12 360 入口 UX", "「360° 全景」按钮、无资源灰显与提示、退出全景", "CompletedStoryViewerRoot", "已更新"],
        # ── VR 功能（底层虚拟/视觉技术）──
        ["功能", "VR", "F-V01 摄像头采集", "WebCamTexture 俯拍摄像头画面", "ArUcoDetector", "已完成"],
        ["功能", "VR", "F-V02 ArUco 检测", "OpenCV 实时识别 marker ID 与像素坐标", "ArUcoDetector", "已完成"],
        ["功能", "VR", "F-V03 AR 贴纸叠加", "在摄像头画面上叠 2D 角色贴纸、名单提示", "StoryCreationArDirector", "已完成"],
        ["功能", "VR", "F-V04 空间布局推断", "从 marker 位置推断角色相对布局", "StoryCreationGapAnalyzer", "已完成"],
        ["功能", "VR", "F-V05 3D 模型预览", "教程中 RenderTexture 展示积木模型、拖拽旋转", "TutorialPreview3DOverlay", "已完成"],
        ["功能", "VR", "F-V06 AI 绘本生图", "img2img 生成故事页插图、角色形象一致", "LocalImageGenClient / Pipeline", "已完成"],
        ["功能", "VR", "F-V07 360° 全景生成", "平面页扩展为 2:1 环视全景图", "generate-panorama API", "部分完成"],
        ["功能", "VR", "F-V08 全景文件存储", "page_XX_panorama.png 与 story.json 关联", "CompletedStoryStore", "已完成"],
        ["功能", "VR", "F-V09 360° 播放器", "equirectangular 贴球面、陀螺仪/鼠标环视", "GyroPanorama360Player", "部分完成"],
        ["功能", "VR", "F-V10 全景切换逻辑", "平面↔360 切换、翻页时重载全景", "CompletedStoryViewerRoot", "已更新"],
        ["功能", "VR", "F-V11 生图网关", "Node.js 对接百炼 wan2.6 文生图/img2img/全景", "storybricks-image-gen-web", "已完成"],
    ]
    write_phase_sheet(ws4, "阶段三：功能实现（每项仅归属一门课，不重复）", ["阶段", "课程", "功能编号", "功能说明", "实现位置", "状态"], func_rows)
    set_col_widths(ws4, [8, 8, 18, 36, 28, 10])

    # ═══════════════════════════════════════
    # Sheet 5：交界说明（避免两课写同一件事）
    # ═══════════════════════════════════════
    ws5 = wb.create_sheet("4-交界不重复")
    ws5.append(["交界话题", "HCI 怎么写（体验侧）", "VR 怎么写（技术侧）", "绝对不重复的内容"])
    style_header(ws5)
    boundary = [
        ["创作页摄像头", "小窗可展开，方便确认积木摆放", "WebCam 采集帧率、ArUco 检测算法、贴纸坐标公式", "HCI 不写 OpenCV；VR 不写按钮位置"],
        ["角色识别反馈", "乐乐说「兔子到了！」、名单提示文案", "marker ID→Sprite 映射、CharacterArrived 事件触发", "HCI 不写 marker 编号规则"],
        ["AI 生成等待", "进度文案、可重新生成、失败提示", "HTTP 请求、img2img JSON、参考图压缩、轮询", "HCI 不写 API；VR 不写按钮文案"],
        ["360° 全景", "按钮在哪、无全景时怎么提示、环视操作说明", "generate-panorama 参数、2:1 贴图、球面 Shader、陀螺仪", "HCI 不写 equirectangular；VR 不写录音功能"],
        ["教程 3D 预览", "「点这里看整模、拖拽旋转」", "RenderTexture 尺寸、Camera 配置、旋转矩阵", "HCI 不写渲染细节"],
        ["故事保存", "保存成功后去「我的故事」回看", "JSON 结构、PNG 路径、panorama 字段", "HCI 不写文件格式"],
        ["引言/背景", "从儿童教育+交互创新切入", "从虚实融合+计算机视觉+生成式 AI 切入", "背景段各写各的问题，不复制"],
    ]
    for i, row in enumerate(boundary, 2):
        for c, val in enumerate(row, 1):
            cell = ws5.cell(row=i, column=c, value=val)
            cell.alignment = WRAP
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    set_col_widths(ws5, [16, 32, 32, 28])

    # ═══════════════════════════════════════
    # Sheet 6：测试
    # ═══════════════════════════════════════
    ws6 = wb.create_sheet("5-测试")
    test_rows = [
        ["测试", "HCI", "T-H01 创作任务完成率", "被试能否独立完成：选故事→搭积木→摆场景→回答乐乐→生成保存", "3~5 人", "≥80%", "任务观察表", "待做"],
        ["测试", "HCI", "T-H02 各阶段耗时", "记录每阶段用时，找瓶颈步骤", "同上", "—", "计时表", "待做"],
        ["测试", "HCI", "T-H03 错误与求助", "卡住几次、是否需要成人帮助", "同上", "—", "观察记录", "待做"],
        ["测试", "HCI", "T-H04 主观满意度", "SUS 或 5 点 Likert 问卷", "同上", "≥4/5", "问卷", "待做"],
        ["测试", "HCI", "T-H05 语音交互理解度", "是否听懂乐乐的提示与追问", "同上", "≥4/5", "问卷 1~2 题", "待做"],
        ["测试", "HCI", "T-H06 360 按钮可发现性", "无提示情况下能否找到全景入口", "同上", "≥80%", "任务观察", "待做"],
        ["测试", "VR", "T-V01 ArUco 识别率", "固定摆放重复检测 20 次", "1 人操作", "≥90%", "计数表", "待做"],
        ["测试", "VR", "T-V02 贴纸对齐主观评价", "贴纸是否跟随 marker 移动", "—", "可用", "截图", "待做"],
        ["测试", "VR", "T-V03 AI 生图成功率", "连续生成 10 页绘本", "—", "≥8/10", "日志", "待做"],
        ["测试", "VR", "T-V04 全景生成成功率", "连续生成 10 页 360", "—", "≥6/10", "日志", "待做"],
        ["测试", "VR", "T-V05 2D vs 360 沉浸感", "同一页分别平面读和 360 环视，5 题问卷", "5~8 人", "360>2D", "问卷", "待做"],
        ["测试", "VR", "T-V06 360 环视舒适度", "360 模式持续 3 分钟", "同上", "≥3/5", "问卷", "待做"],
    ]
    write_phase_sheet(ws6, "阶段四：测试（两课指标完全分开）", ["阶段", "课程", "测试项", "方法", "被试", "通过标准", "记录", "状态"], test_rows)
    set_col_widths(ws6, [8, 8, 22, 28, 10, 10, 12, 8])

    # ═══════════════════════════════════════
    # Sheet 7：总结 + 报告目录
    # ═══════════════════════════════════════
    ws7 = wb.create_sheet("6-总结与报告目录")
    ws7.append(["课程", "报告建议目录", "对应阶段/功能", "页数", "状态"])
    style_header(ws7)
    outline = [
        ["HCI", "1 引言（儿童创作交互问题）", "策划", "1~2", "待写"],
        ["HCI", "2 需求分析与用户研究", "策划", "2~3", "待写"],
        ["HCI", "3 信息架构与流程设计", "设计", "3~4", "待写"],
        ["HCI", "4 交互设计（状态机/语音/教程/阅读）", "设计", "4~5", "待写"],
        ["HCI", "5 界面设计", "设计", "2~3", "待写"],
        ["HCI", "6 功能实现概述（F-H01~H12）", "功能", "2~3", "待写"],
        ["HCI", "7 可用性测试（T-H01~H06）", "测试", "3~4", "待做"],
        ["HCI", "8 总结与展望", "总结", "1", "待写"],
        ["VR", "1 引言（虚实融合与 360 呈现问题）", "策划", "1~2", "待写"],
        ["VR", "2 相关技术与可行性分析", "策划", "2~3", "待写"],
        ["VR", "3 系统架构与方案设计", "设计", "2~3", "待写"],
        ["VR", "4 虚实融合实现（F-V01~V04）", "功能", "3~4", "待写"],
        ["VR", "5 三维预览与 AI 生图（F-V05~V06,V11）", "功能", "3~4", "待写"],
        ["VR", "6 360° 全景管线与播放（F-V07~V10）", "功能", "3~4", "待完善"],
        ["VR", "7 测试与评估（T-V01~V06）", "测试", "3~4", "待做"],
        ["VR", "8 总结与展望", "总结", "1", "待写"],
    ]
    for i, row in enumerate(outline, 2):
        for c, val in enumerate(row, 1):
            cell = ws7.cell(row=i, column=c, value=val)
            cell.alignment = WRAP
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if c == 1:
                cell.fill = HCI_FILL if val == "HCI" else VR_FILL
    set_col_widths(ws7, [8, 36, 10, 8, 8])

    # ═══════════════════════════════════════
    # Sheet 8：时间线（按阶段对齐）
    # ═══════════════════════════════════════
    ws8 = wb.create_sheet("7-时间线")
    ws8.append(["周次", "阶段", "HCI 工作", "VR 工作", "阶段交付物"])
    style_header(ws8)
    timeline = [
        ["第1周", "策划", "用户画像、任务分析、竞品", "可行性、标记方案、内容规格、技术栈", "各写需求/可行性章节"],
        ["第2周", "设计", "IA、状态机、语音脚本、UI 线框", "AR 方案、生图管线、360 管线设计", "设计图+流程图"],
        ["第3周", "功能", "教程/创作/阅读 UI 整理与截图", "ArUco/AR/生图/全景代码整理", "功能清单对照"],
        ["第4周", "功能", "联调创作流程、360 按钮 UX", "P-01~P-04 全景补全", "端到端 demo"],
        ["第5周", "测试", "T-H01~H06 可用性测试", "T-V01~V04 技术指标测试", "原始数据"],
        ["第6周", "测试", "整理 HCI 测试结论", "T-V05~V06 沉浸感对比", "图表"],
        ["第7周", "总结", "HCI 报告初稿", "VR 报告初稿", "互引 1 段"],
        ["第8周", "总结", "定稿+答辩 PPT", "定稿+答辩 PPT", "HCI demo 创作 / VR demo 360"],
    ]
    for i, row in enumerate(timeline, 2):
        for c, val in enumerate(row, 1):
            cell = ws8.cell(row=i, column=c, value=val)
            cell.alignment = WRAP
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if c == 2 and val in PHASE_FILLS:
                cell.fill = PHASE_FILLS[val]
    set_col_widths(ws8, [8, 8, 28, 28, 22])

    # ═══════════════════════════════════════
    # Sheet 9：全景待完善（VR 课）
    # ═══════════════════════════════════════
    ws9 = wb.create_sheet("8-全景待完善")
    ws9.append(["编号", "任务", "方案", "工时", "归属", "状态"])
    style_header(ws9)
    pano = [
        ["P-01", "统一全景文件命名", "page_XX_panorama.jpg，兼容 _pano 旧名", "0.5天", "VR/功能", "待做"],
        ["P-02", "生成进度与失败提示", "「正在生成 360°…」+ 失败 Toast", "1天", "VR 技术 / HCI 只看文案", "待做"],
        ["P-03", "360 按钮 UX", "无全景灰显+提示", "0.5天", "HCI/功能", "已完成"],
        ["P-04", "回正视角", "ResetLook 按钮", "0.5天", "VR/功能", "待做"],
        ["P-05", "StreamingAssets 演示", "无 AI 时 fallback 演示全景", "1天", "VR/功能", "待做"],
        ["P-06", "翻页平滑切换", "预加载减黑屏", "1天", "VR/功能", "待做"],
        ["P-07", "Editor 演示菜单", "一键绑定龟兔赛跑全景", "0.5天", "VR/功能", "待做"],
        ["P-08", "全景 prompt 优化", "无缝环视、地平线一致", "0.5天", "VR/功能", "待做"],
    ]
    for i, row in enumerate(pano, 2):
        for c, val in enumerate(row, 1):
            cell = ws9.cell(row=i, column=c, value=val)
            cell.alignment = WRAP
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    set_col_widths(ws9, [8, 22, 36, 8, 16, 8])

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
